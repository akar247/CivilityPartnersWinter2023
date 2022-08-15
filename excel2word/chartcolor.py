from openpyxl import load_workbook
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.label import DataLabelList

wb = load_workbook(filename="Survey.xlsx") 
ws = wb.active
#workbook.sheetnames
#workbook.worksheets
import docx





for ws in wb.worksheets:
        
        for row in ws.iter_rows(min_col=1, max_col=1):
            if row[0].value is None or row[0].value == "Reason":
                maxrowchart = (row[0].row - 1)
                break
        values = Reference(ws, min_col= 2, min_row=3, max_col = 2, max_row = maxrowchart)
        cats = Reference(ws, min_col=1, max_col=1, min_row=4, max_row= maxrowchart)

                # Create object of BarChart class
        chart = BarChart()
        chart.height = 10
        chart.width = 15
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)
                # set the title of the chart
        chart.title = ws["A2"].value

                # the top-left corner of the chart
                # is anchored to cell F2 .
        chart.varyColors = "000F0FFF"
        
        for i in range(1,50):
                s=ws[f"A{i}"].value
                if s == 'Answer Choices': 
                        ws.add_chart(chart,"H2")
                
                        
        
wb.save("test.xlsx")


from win32com.client import Dispatch

def chart2():
    app = Dispatch("Excel.Application")
    workbook_file_name = r"C:\Users\damon\OneDrive\Documents\GitHub\CPxTCG\excel2word\test.xlsx"
    workbook = app.Workbooks.Open(Filename=workbook_file_name)

    app.DisplayAlerts = False

    i = 1
    for sheet in workbook.Worksheets:
        for chartObject in sheet.ChartObjects():
            print(sheet.Name + ':' + chartObject.Name)
            chartObject.Chart.Export(r"C:\Users\damon\OneDrive\Documents\GitHub\CPxTCG\excel2word\Images\Survey" + sheet.Name + ".png")
            i += 1

    workbook.Close(SaveChanges=False, Filename=workbook_file_name)

chart2()
        
        

